import json
import csv
import os
import io
import re
from io import StringIO
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Alignment, Font

from flask import Blueprint, request, send_file

from hanfor_flask import current_app, nocache, HanforFlask
from lib_core.utils import get_requirements, generate_file_response, generate_req_file_content
from lib_core.data import VariableCollection, SessionValue, Tag

api_blueprint = Blueprint("tools_api", __name__, url_prefix="/api/tools")


@api_blueprint.route("/<command>", methods=["GET", "POST"])
@nocache
def tools_api(command):
    filter_list = request.form.get("selected_requirement_ids", "")
    if len(filter_list) > 0:
        filter_list = json.loads(filter_list)
    else:
        filter_list = None

    file_name = f"{current_app.config['CSV_INPUT_FILE'][:-4]}"

    if command == "req_file":
        content = generate_req_file_content(current_app, filter_list=filter_list)
        return generate_file_response(content, file_name + ".req")

    if command == "csv_file":
        file = generate_csv_file_content(current_app, filter_list=filter_list)
        name = f"{current_app.config['SESSION_TAG']}_{current_app.config['USING_REVISION']}_out.csv"
        return generate_file_response(file, name, mimetype="text/csv")

    if command == "xls_file":
        file = generate_xls_file_content(current_app, filter_list=filter_list)
        file.seek(0)
        return send_file(file, download_name=file_name + ".xlsx", as_attachment=True)


def generate_csv_file_content(app: HanforFlask, filter_list=None, invert_filter=False):  # TODO nach tools
    """Generates the csv file content for a session.

    :param app: Current hanfor app for context.
    :type app: Flaskapp
    :param filter_list: (Optional) A list of requirement IDs to be included in the result. All if not set.
    :type filter_list: list (of strings)
    :param invert_filter: Exclude filter
    :type invert_filter: bool
    :return: CSV content
    :rtype: str
    """
    # Get requirements
    requirements = get_requirements(app, filter_list=filter_list, invert_filter=invert_filter)

    # get session status
    session_values = app.db.get_objects(SessionValue)

    # Add Formalization col if not existent in input CSV.
    for csv_key in [session_values["csv_formal_header"].value]:
        if csv_key not in session_values["csv_fieldnames"].value:
            session_values["csv_fieldnames"].value.append(csv_key)
        app.db.update()

    # Add Hanfor Tag col to csv.
    # TODO: remove static column names and replace with config via user startup config dialog.
    tag_col_name = "Hanfor_Tags"
    status_col_name = "Hanfor_Status"
    for col_name in [tag_col_name, status_col_name]:
        if col_name not in session_values["csv_fieldnames"].value:
            session_values["csv_fieldnames"].value.append(col_name)
        app.db.update()

    # Update csv row of requirements to use their latest formalization and tags.
    for requirement in requirements:
        requirement.csv_row[session_values["csv_formal_header"].value] = requirement.get_formalizations_json()
        requirement.csv_row[tag_col_name] = ", ".join(requirement.get_tag_name_comment_dict().keys())
        requirement.csv_row[status_col_name] = requirement.status

    # Write data to file.
    rows = [r.csv_row for r in requirements]
    with StringIO() as out_csv:
        csv.register_dialect("ultimate", delimiter=",")
        writer = csv.DictWriter(out_csv, session_values["csv_fieldnames"].value)
        writer.writeheader()
        writer.writerows(rows)
        result = out_csv.getvalue()

    return result


def generate_xls_file_content(
    app: HanforFlask, filter_list: list[str] = None, invert_filter: bool = False
) -> io.BytesIO:  # TODO nach tools
    """Generates the xlsx file content for a session."""

    requirements = get_requirements(app, filter_list=filter_list, invert_filter=invert_filter)
    var_collection = VariableCollection(app)
    tags = {tag.name: tag.internal for tag in app.db.get_objects(Tag).values()}

    # create  styles
    style_multiline = Alignment(vertical="top", wrap_text=True)
    style_bold = Font(bold=True)
    style_white = Font(color="FFFFFF", bold=True)
    style_filled = PatternFill(fill_type="solid", start_color="2a6ebb", end_color="2a6ebb")
    style_meta = PatternFill(fill_type="solid", start_color="004a99", end_color="004a99")
    # create excel template
    work_book = Workbook()
    work_sheet = work_book.active
    work_sheet.title = "Requirements"

    header_offset = 4

    def make_header(ws):
        ws.freeze_panes = "A4"
        for col in range(1, 10):
            for r in range(1, 3):
                ws.cell(r, col).fill = style_meta
        ws.cell(1, 2, value="HANFOR Report")
        ws.cell(1, 2).font = style_white
        ws.cell(1, 3, value=os.path.basename(app.db.get_object(SessionValue, "csv_input_file").value))
        ws.cell(1, 3).font = Font(color="FFFFFF")
        for col in range(1, 10):
            ws.cell(header_offset - 1, col).fill = style_filled
            ws.cell(header_offset - 1, col).font = style_white

    make_header(work_sheet)

    # Set column widths and headings
    work_sheet.column_dimensions["A"].width = 5
    work_sheet.cell(header_offset - 1, 1, value="Index")
    work_sheet.column_dimensions["B"].width = 20
    work_sheet.cell(header_offset - 1, 2, value="ID")
    work_sheet.column_dimensions["C"].width = 80
    work_sheet.cell(header_offset - 1, 3, value="Description")
    work_sheet.cell(header_offset - 1, 4, value="Type")
    work_sheet.column_dimensions["E"].width = 40
    work_sheet.cell(header_offset - 1, 5, value="Tags")
    work_sheet.cell(header_offset - 1, 6, value="Status")
    work_sheet.column_dimensions["G"].width = 160
    work_sheet.cell(header_offset - 1, 7, value="Formalisation")

    for i, requirement in enumerate(requirements):
        for c in range(1, 8):
            # Note: setting styles is ordering-sensitive so set styles FIRST
            work_sheet.cell(header_offset + i, c).alignment = style_multiline
        work_sheet.cell(header_offset + i, 1, requirement.pos_in_csv)
        work_sheet.cell(header_offset + i, 2).font = style_bold
        work_sheet.cell(header_offset + i, 2, requirement.rid)
        work_sheet.cell(header_offset + i, 3, requirement.description)
        work_sheet.cell(header_offset + i, 4, requirement.type_in_csv)

        work_sheet.cell(
            header_offset + i,
            5,
            "".join(
                [
                    f"{t}: {c} \n" if c else f"{t}\n"
                    for t, c in requirement.get_tag_name_comment_dict().items()
                    if t in tags and not tags[t]
                ]
            ),
        )
        work_sheet.cell(header_offset + i, 6, requirement.status)
        work_sheet.cell(header_offset + i, 7, "\n".join([f.get_string() for f in requirement.formalizations.values()]))

    # make severity sheet
    tag_sheet = work_book.create_sheet("Findings")
    make_header(tag_sheet)
    tag_sheet.column_dimensions["A"].width = 5
    tag_sheet.cell(header_offset - 1, 1, value="Index")
    tag_sheet.column_dimensions["B"].width = 20
    tag_sheet.cell(header_offset - 1, 2, value="ID")
    tag_sheet.cell(header_offset - 1, 3, value="Description")
    tag_sheet.column_dimensions["C"].width = 80
    tag_sheet.column_dimensions["D"].width = 20
    tag_sheet.cell(header_offset - 1, 4, value="Tag")
    tag_sheet.column_dimensions["E"].width = 60
    tag_sheet.cell(header_offset - 1, 5, value="Comment (Analysis)")
    tag_sheet.column_dimensions["F"].width = 10
    tag_sheet.cell(header_offset - 1, 6, value="Accept")
    tag_sheet.column_dimensions["G"].width = 15
    tag_sheet.cell(header_offset - 1, 7, value="Value")
    tag_sheet.column_dimensions["H"].width = 80
    tag_sheet.cell(header_offset - 1, 8, value="Comment (Review)")

    accept_state_validator = DataValidation(
        type="list",
        formula1='"TODO,'
        "Accept,"
        "Bycatch (wrong report; defective anyways),"
        "Decline,"
        "Inquery (report is unclear),"
        'Undecided (investigation is needed)"',
        allow_blank=False,
    )
    tag_sheet.add_data_validation(accept_state_validator)
    accept_state_validator.add("F4:F1048576")
    issue_value_validator = DataValidation(
        type="list", formula1='"TODO, 0 (no value),1 (nice to have),2 (useful),3 (possible desaster)"', allow_blank=True
    )
    tag_sheet.add_data_validation(issue_value_validator)
    issue_value_validator.add("G4:G1048576")

    issue_tags_reqs = []
    for req in requirements:
        for tag in req.get_tag_name_comment_dict():
            if tag in tags and tags[tag]:
                continue
            issue_tags_reqs.append((req, tag))

    for i, (req, tag) in enumerate(issue_tags_reqs):
        for c in range(1, 8):
            tag_sheet.cell(header_offset + i, c).alignment = style_multiline
        tag_sheet.cell(header_offset + i, 1, req.pos_in_csv)
        tag_sheet.cell(header_offset + i, 2).font = style_bold
        tag_sheet.cell(header_offset + i, 2, req.rid)
        tag_sheet.cell(header_offset + i, 3, req.description)
        tag_sheet.cell(header_offset + i, 4, tag)
        tag_sheet.cell(
            header_offset + i, 5, req.get_tag_name_comment_dict()[tag]
        )  # Tags do currently not have comments
        tag_sheet.cell(header_offset + i, 6, "TODO")
        tag_sheet.cell(header_offset + i, 7, "TODO")

    # make sheet with variables
    var_sheet = work_book.create_sheet("Variables")
    make_header(var_sheet)
    var_sheet.column_dimensions["A"].width = 40
    var_sheet.cell(header_offset - 1, 1, value="Name")
    var_sheet.column_dimensions["B"].width = 80
    var_sheet.column_dimensions["C"].width = 5
    var_sheet.cell(header_offset - 1, 2, value="Type")
    var_sheet.column_dimensions["D"].width = 180
    var_sheet.cell(header_offset - 1, 4, value="Invarianten")

    for i, var in enumerate(var_collection.collection.values()):
        for c in range(1, 8):
            var_sheet.cell(header_offset + i, c).alignment = style_multiline
        var_sheet.cell(header_offset + i, 1, var.name)
        var_sheet.cell(header_offset + i, 1).font = style_bold
        var_sheet.cell(header_offset + i, 2, var.type)
        var_sheet.cell(header_offset + i, 3, "E" if var.belongs_to_enum else "")
        var_sheet.cell(header_offset + i, 4, "\n".join([c.get_string() for c in var.get_constraints().values()]))

    work_book.active = tag_sheet
    buffer = io.BytesIO()
    work_book.save(buffer)
    return buffer


def clean_identifier_for_ultimate_parser(req_id: str, formalisation_id: int, used_identifiers: set[str]) -> str:
    """Clean slug to be sound for ultimate parser.

    :param req_id: The requirement id to be cleaned.
    :param formalisation_id: The formalisation id to be cleaned.
    :param used_identifiers: Set of already used identifiers.
    :return: (identifier, used_slugs) save_slug a save to use form of slug. save_slug added to used_slugs.
    """
    # Replace any occurrence of [whitespace, `.`, `-`] with `_`
    base_identifier = re.sub(r"[\s+.-]+", "_", req_id.strip())

    # Resolve illegal start by prepending the slug with ID_ in case it does not start with a letter.
    base_identifier = re.sub(r"^([^a-zA-Z])", r"ID_\1", base_identifier)

    def create_identifier(base: str, extension: int, base_suffix: int = -1):
        if base_suffix == -1:
            return "{}_{}".format(base, extension)
        return "{}_{}_{}".format(base, base_suffix, extension)

    identifier = create_identifier(base_identifier, formalisation_id)

    # Resolve duplicates
    # search for the first free suffix.
    if identifier in used_identifiers:
        suffix = 1

        while create_identifier(base_identifier, formalisation_id, suffix) in used_identifiers:
            suffix += 1
        identifier = create_identifier(base_identifier, formalisation_id, suffix)

    used_identifiers.add(identifier)

    return identifier
