""" 

@copyright: 2018 Samuel Roth <samuel@smel.de>
@licence: GPLv3
"""

import argparse
import io
import shutil
from collections import defaultdict
from io import StringIO

import csv
import datetime
import html
import logging
import os
from shutil import copytree
import re
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Alignment, Font

from hanfor_flask import HanforFlask
from flask import json, Response
from flask_assets import Environment
from patterns import PATTERNS

import boogie_parsing

from json_db_connector.json_db import DatabaseTable, TableType, DatabaseID, DatabaseField, JsonDatabase, remove_json_database_data_tracing_logger
from dataclasses import dataclass, field
from uuid import uuid4
from typing import Callable
from tags.tags import Tag
from defaults import Color

# Here is the first time we use config. Check existence and raise a meaningful exception if not found.
try:
    from config import PATTERNS_GROUP_ORDER
except ModuleNotFoundError:
    msg = "Missing a config file. See README.md -> #Setup on how to create one."
    logging.error(msg)
    raise FileNotFoundError(msg)

from reqtransformer import VariableCollection, Requirement, RequirementCollection, Variable, Scope
from static_utils import (
    replace_prefix,
    hash_file_sha1,
    SessionValue,
)
from typing import Union, Set, List
from terminaltables import DoubleTable

here = os.path.dirname(os.path.realpath(__file__))
default_scope_options = """
    <option value="NONE">None</option>
    <option value="GLOBALLY">Globally</option>
    <option value="BEFORE">Before "{P}"</option>
    <option value="AFTER">After "{P}"</option>
    <option value="BETWEEN">Between "{P}" and "{Q}"</option>
    <option value="AFTER_UNTIL">After "{P}" until "{Q}"</option>
    """


@DatabaseTable(TableType.File)
@DatabaseID("id", use_uuid=True)
@DatabaseField("timestamp", datetime.datetime)
@DatabaseField("message", str)
@DatabaseField("requirements", list[Requirement])
@dataclass()
class RequirementEditHistory:
    timestamp: datetime.datetime
    message: str
    requirements: list[Requirement]
    id: str = field(default_factory=lambda: str(uuid4()))


def config_check(app_config):
    to_ensure_configs = ["PATTERNS_GROUP_ORDER"]
    for to_ensure_config in to_ensure_configs:
        if to_ensure_config not in app_config:
            raise SyntaxError("Could not find {} in config.".format(to_ensure_config))

    # Check pattern groups set correctly.
    pattern_groups_used = set((pattern["group"] for pattern in PATTERNS.values()))
    pattern_groups_set = set((group for group in app_config["PATTERNS_GROUP_ORDER"]))
    # Pattern group for downwards compatibility: Legacy patterns are not shown in dropdown (but still deserializable)
    pattern_groups_set.add("Legacy")

    if not pattern_groups_used == pattern_groups_set:
        if len(pattern_groups_used - pattern_groups_set) > 0:
            raise SyntaxError(
                "No group order set in config for pattern groups {}".format(pattern_groups_used - pattern_groups_set)
            )

    try:
        get_default_pattern_options()
    except Exception as e:
        raise SyntaxError("Could not parse pattern config. Please check your config.py: {}.".format(e))


def get_default_pattern_options():
    """Parse the pattern config into the dropdown list options for the frontend

    Returns (str): Options for pattern selection in HTML.

    """
    result = '<option value="NotFormalizable">None</option>'
    opt_group_lists = defaultdict(list)
    opt_groups = defaultdict(str)
    # Collect pattern in groups.
    for name, pattern_dict in PATTERNS.items():
        opt_group_lists[pattern_dict["group"]].append((pattern_dict["pattern_order"], name, pattern_dict["pattern"]))

    # Sort groups and concatenate pattern options
    for group_name, opt_list in opt_group_lists.items():
        for _, name, pattern in sorted(opt_list):
            option = (
                '<option value="{name}">{pattern}</option>'.format(name=name, pattern=pattern)
                .replace("{", '"{')
                .replace("}", '}"')
            )
            opt_groups[group_name] += option

    # Enclose pattern options by their groups.
    for group_name in PATTERNS_GROUP_ORDER:
        result += '<optgroup label="{}">'.format(group_name)
        result += opt_groups[group_name]
        result += "</optgroup>"

    return result


def get_formalization_template(templates_folder, formalization_id, formalization):
    result = {
        "success": True,
        "html": formalization_html(
            templates_folder, formalization_id, default_scope_options, get_default_pattern_options(), formalization
        ),
    }

    return result


def formalization_html(templates_folder, formalization_id, scope_options, pattern_options, formalization):
    # Load template.
    html_template = ""
    with open(os.path.join(templates_folder, "formalization.html"), mode="r") as f:
        html_template += "\n".join(f.readlines())

    # Set values
    html_template = html_template.replace("__formalization_text__", formalization.get_string())
    html_template = html_template.replace("__formal_id__", "{}".format(formalization_id))
    form_desc = formalization.get_string()
    if len(form_desc) < 10:  # Add hint to open if desc is short.
        form_desc += "... (click to open)"
    html_template = html_template.replace("__formal_desc__", form_desc)

    # Selected scope and pattern:
    if formalization.scoped_pattern is not None:
        scope = formalization.scoped_pattern.get_scope_slug()
        pattern = formalization.scoped_pattern.get_pattern_slug()
        scope_options = scope_options.replace('value="{}"'.format(scope), 'value="{}" selected'.format(scope))
        pattern_options = pattern_options.replace('value="{}"'.format(pattern), 'value="{}" selected'.format(pattern))
    html_template = html_template.replace("__scope_options__", scope_options)
    html_template = html_template.replace("__pattern__options__", pattern_options)

    # Expressions
    for key, value in formalization.expressions_mapping.items():
        html_template = html_template.replace(
            "__expr_{}_content__".format(key), "{}".format(html.escape(str(value.raw_expression)))
        )

    # Unset remaining vars.
    html_template = re.sub(r"__expr_._content__", "", html_template)

    return html_template


def formalizations_to_html(app: HanforFlask, formalizations):
    result = ""
    for index, formalization in formalizations.items():
        result += formalization_html(
            app.config["TEMPLATES_FOLDER"], index, default_scope_options, get_default_pattern_options(), formalization
        )
    return result


def get_available_vars(app: HanforFlask, full=True):
    var_collection = VariableCollection(app)
    result = var_collection.get_available_vars_list(used_only=not full)
    return result


def get_requirements_using_var(requirements: list, var_name: str):
    """Return a list of requirement ids where var_name is used in at least one formalization.

    :param requirements: list of Requirement.
    :param var_name: Variable name to search for.
    :return: List of affected Requirement ids.
    """
    result_rids = []
    for requirement in requirements:  # type: Requirement
        if requirement.uses_var(var_name):
            result_rids.append(requirement.rid)

    return result_rids


def update_variable_in_collection(app: HanforFlask, request):
    """Update a single variable. The request should contain a form:
        name -> the new name of the var.
        name_old -> the name of the var before.
        type -> the new type of the var.
        type_old -> the old type of the var.
        occurrences -> Ids of requirements using this variable.
        enumerators -> The dict of enumerators

    :param app: the running flask app
    :type app: HanforFlask
    :param request: A form request
    :type request: flask.Request
    :return: Dictionary containing changed data and request status information.
    :rtype: dict
    """
    # Get properties from request
    var_name = request.form.get("name", "").strip()
    var_name_old = request.form.get("name_old", "").strip()
    var_type = request.form.get("type", "").strip()
    var_type_old = request.form.get("type_old", "").strip()
    var_const_val = request.form.get("const_val", "").strip()
    var_const_val_old = request.form.get("const_val_old", "").strip()
    belongs_to_enum = request.form.get("belongs_to_enum", "").strip()
    belongs_to_enum_old = request.form.get("belongs_to_enum_old", "").strip()
    occurrences = request.form.get("occurrences", "").strip().split(",")
    enumerators = json.loads(request.form.get("enumerators", ""))

    # TODO: remove
    while "" in occurrences:
        occurrences.remove("")

    var_collection = VariableCollection(app)
    result = {
        "success": True,
        "has_changes": False,
        "type_changed": False,
        "name_changed": False,
        "rebuild_table": False,
        "data": {"name": var_name, "type": var_type, "used_by": occurrences, "const_val": var_const_val},
    }

    # Check for changes
    if (
        var_type_old != var_type
        or var_name_old != var_name
        or var_const_val_old != var_const_val
        or request.form.get("updated_constraints") == "true"
        or belongs_to_enum != belongs_to_enum_old
    ):
        logging.info("Update Variable `{}`".format(var_name_old))
        result["has_changes"] = True
        reload_type_inference = False

        # Update type.
        if var_type_old != var_type:
            logging.info("Change type from `{}` to `{}`.".format(var_type_old, var_type))
            try:
                var_collection.collection[var_name_old].belongs_to_enum = belongs_to_enum
                var_collection.set_type(var_name_old, var_type)
            except TypeError as e:
                result = {"success": False, "errormsg": str(e)}
                return result
            result["type_changed"] = True
            reload_type_inference = True

        # Update const value.
        if var_const_val_old != var_const_val:
            try:
                if var_type == "ENUMERATOR_INT":
                    int(var_const_val)
                if var_type in ["ENUMERATOR_REAL"]:
                    float(var_const_val)
            except Exception as e:
                result = {
                    "success": False,
                    "errormsg": "Enumerator value `{}` for {} `{}` not valid: {}".format(
                        var_const_val, var_type, var_name, e
                    ),
                }
                return result
            logging.info("Change value from `{}` to `{}`.".format(var_const_val_old, var_const_val))
            var_collection.collection[var_name].value = var_const_val
            result["val_changed"] = True

        # Update constraints.
        if request.form.get("updated_constraints") == "true":
            constraints = json.loads(request.form.get("constraints", ""))
            logging.debug("Update Variable Constraints")
            try:
                var_collection = var_collection.collection[var_name_old].update_constraints(
                    constraints, app, var_collection
                )
                result["rebuild_table"] = True
                app.db.update()
            except KeyError as e:
                result["success"] = False
                result["error_msg"] = "Could not set constraint: Missing expression/variable for {}".format(e)
            except Exception as e:
                result["success"] = False
                result["error_msg"] = "Could not parse formalization: `{}`".format(e)
        else:
            logging.debug("Skipping variable Constraints update.")

        # update name.
        if var_name_old != var_name:
            logging.debug("Change name of var `{}` to `{}`".format(var_name_old, var_name))
            #  Case: New name which does not exist -> remove the old var and replace in reqs occurring.
            if var_name not in var_collection:
                logging.debug("`{}` is a new var name. Rename the var, replace occurrences.".format(var_name))
                # Rename the var (Copy to new name and remove the old item. Rename it)
                affected_enumerators = var_collection.rename(var_name_old, var_name, app)

            else:  # Case: New name exists. -> Merge the two vars into one. -> Complete rebuild.
                if var_collection.collection[var_name_old].type != var_collection.collection[var_name].type:
                    result["success"] = False
                    result["errormsg"] = "To merge two variables the types must be identical. {} != {}".format(
                        var_collection.collection[var_name_old].type, var_collection.collection[var_name].type
                    )
                    return result
                logging.debug("`{}` is an existing var name. Merging the two vars. ".format(var_name))
                # var_collection.merge_vars(var_name_old, var_name, app)
                affected_enumerators = var_collection.rename(var_name_old, var_name, app)
                result["rebuild_table"] = True
                reload_type_inference = True

            # Update the requirements using this var.
            if len(occurrences) > 0:
                rename_variable_in_expressions(app, occurrences, var_name_old, var_name)

            for old_enumerator_name, new_enumerator_name in affected_enumerators:
                # Todo: Implement this more eff.
                # Todo: Refactor Formalizations to use hashes as vars and fetch them in __str__ from the var collection.
                requirements = get_requirements(app)
                affected_requirements = get_requirements_using_var(requirements, old_enumerator_name)
                rename_variable_in_expressions(app, affected_requirements, old_enumerator_name, new_enumerator_name)

            result["name_changed"] = True

        # Change ENUM parent.
        if belongs_to_enum != belongs_to_enum_old and var_type in ["ENUMERATOR_INT", "ENUMERATOR_REAL"]:
            logging.debug("Change enum parent of enumerator `{}` to `{}`".format(var_name, belongs_to_enum))
            if belongs_to_enum not in var_collection:
                result["success"] = False
                result["errormsg"] = "The new ENUM parent `{}` does not exist.".format(belongs_to_enum)
                return result
            if var_collection.collection[belongs_to_enum].type != replace_prefix(var_type, "ENUMERATOR", "ENUM"):
                result["success"] = False
                result["errormsg"] = "The new ENUM parent `{}` is not an {} (is `{}`).".format(
                    belongs_to_enum,
                    replace_prefix(var_type, "ENUMERATOR", "ENUM"),
                    var_collection.collection[belongs_to_enum].type,
                )
                return result
            new_enumerator_name = replace_prefix(var_name, belongs_to_enum_old, "")
            new_enumerator_name = replace_prefix(new_enumerator_name, "_", "")
            new_enumerator_name = replace_prefix(new_enumerator_name, "", belongs_to_enum + "_")
            if new_enumerator_name in var_collection:
                result["success"] = False
                result["errormsg"] = "The new ENUM parent `{}` already has a ENUMERATOR `{}`.".format(
                    belongs_to_enum, new_enumerator_name
                )
                return result

            var_collection.collection[var_name].belongs_to_enum = belongs_to_enum
            var_collection.rename(var_name, new_enumerator_name, app)

        logging.info("Store updated variables.")
        var_collection.refresh_var_constraint_mapping()
        var_collection.store()
        app.db.update()
        logging.info("Update derived types by parsing affected formalizations.")
        if reload_type_inference and var_name in var_collection.var_req_mapping:
            for rid in var_collection.var_req_mapping[var_name]:
                if app.db.key_in_table(Requirement, rid):
                    requirement = app.db.get_object(Requirement, rid)
                    requirement.run_type_checks(var_collection)
            app.db.update()

    if var_type in ["ENUM_INT", "ENUM_REAL"]:
        new_enumerators = []
        for enumerator_name, enumerator_value in enumerators:
            if len(enumerator_name) == 0 and len(enumerator_value) == 0:
                continue
            if len(enumerator_name) == 0 or not re.match("^[a-zA-Z0-9_-]+$", enumerator_name):
                result = {"success": False, "errormsg": "Enumerator name `{}` not valid".format(enumerator_name)}
                break
            try:
                if var_type == "ENUM_INT":
                    int(enumerator_value)
                if var_type == "ENUM_REAL":
                    float(enumerator_value)
            except Exception as e:
                result = {
                    "success": False,
                    "errormsg": "Enumerator value `{}` for enumerator `{}` not valid: {}".format(
                        enumerator_value, enumerator_name, e
                    ),
                }
                break

            enumerator_name = "{}_{}".format(var_name, enumerator_name)
            # Add new enumerators to the var_collection
            if not var_collection.var_name_exists(enumerator_name):
                var_collection.add_var(enumerator_name)
                new_enumerators.append(enumerator_name)

            var_collection.collection[enumerator_name].set_type(f"ENUMERATOR_{var_type[5:]}")
            var_collection.collection[enumerator_name].value = enumerator_value
            var_collection.collection[enumerator_name].belongs_to_enum = var_name

        var_collection.store()
        app.db.update()

    return result


def rename_variable_in_expressions(app: HanforFlask, occurrences, var_name_old, var_name):
    """Updates (replace) the variables in the requirement expressions.

    :param app: Flask app (used for session).
    :type app: HanforFlask
    :param occurrences: Requirement ids to be taken into account.
    :type occurrences: list (of str)
    :param var_name_old: The current name in the expressions.
    :type var_name_old: str
    :param var_name: The new name.
    :type var_name: str
    """
    logging.debug("Update requirements using old var `{}` to `{}`".format(var_name_old, var_name))
    for rid in occurrences:
        requirement = app.db.get_object(Requirement, rid)
        # replace in every formalization
        for index, formalization in requirement.formalizations.items():
            for key, expression in formalization.expressions_mapping.items():
                if var_name_old not in expression.raw_expression:
                    continue
                new_expression = boogie_parsing.replace_var_in_expression(
                    expression=expression.raw_expression, old_var=var_name_old, new_var=var_name
                )
                requirement.formalizations[index].expressions_mapping[key].raw_expression = new_expression
                requirement.formalizations[index].expressions_mapping[key].used_variables.discard(var_name_old)
                requirement.formalizations[index].expressions_mapping[key].used_variables.add(var_name)
        logging.debug("Updated variables in requirement id: `{}`.".format(requirement.rid))
    app.db.update()


def get_requirements(app: HanforFlask, filter_list=None, invert_filter=False):
    """Load all requirements from session folder and return in a list.
    Orders the requirements based on their position in the CSV used to create the session (pos_in_csv).

    :param app: current app
    :param filter_list: A list of requirement IDs to be included in the result. All if not set.
    :type filter_list: list (of strings)
    :param invert_filter: Exclude filter
    :type invert_filter: bool
    """

    def should_be_in_result(requirement) -> bool:
        if filter_list is None:
            return True
        return (requirement.rid in filter_list) != invert_filter

    requirements = list()
    for req in app.db.get_objects(Requirement).values():
        if should_be_in_result(req):
            logging.debug("Adding {} to results.".format(req.rid))
            requirements.append(req)

    # We want to preserve the order of the generated CSV relative to the origin CSV.
    requirements.sort(key=lambda x: x.pos_in_csv)

    return requirements


def generate_csv_file_content(app: HanforFlask, filter_list=None, invert_filter=False):
    """Generates the csv file content for a session.

    :param app: Current hanfor app for context.
    :type app: Flaskapp
    :param filter_list: (Optional) A list of requirement IDs to be included in the result. All if not set.
    :type filter_list: list (of strings)
    :param invert_filter: Exclude filter
    :type invert_filter: bool
    :return: CSV content
    :rtype: str
    """
    # Get requirements
    requirements = get_requirements(app, filter_list=filter_list, invert_filter=invert_filter)

    # get session status
    session_values = app.db.get_objects(SessionValue)

    # Add Formalization col if not existent in input CSV.
    for csv_key in [session_values["csv_formal_header"].value]:
        if csv_key not in session_values["csv_fieldnames"].value:
            session_values["csv_fieldnames"].value.append(csv_key)
        app.db.update()

    # Add Hanfor Tag col to csv.
    # TODO: remove static column names and replace with config via user startup config dialog.
    tag_col_name = "Hanfor_Tags"
    status_col_name = "Hanfor_Status"
    for col_name in [tag_col_name, status_col_name]:
        if col_name not in session_values["csv_fieldnames"].value:
            session_values["csv_fieldnames"].value.append(col_name)
        app.db.update()

    # Update csv row of requirements to use their latest formalization and tags.
    for requirement in requirements:
        requirement.csv_row[session_values["csv_formal_header"].value] = requirement.get_formalizations_json()
        requirement.csv_row[tag_col_name] = ", ".join(requirement.get_tag_name_comment_dict().keys())
        requirement.csv_row[status_col_name] = requirement.status

    # Write data to file.
    rows = [r.csv_row for r in requirements]
    with StringIO() as out_csv:
        csv.register_dialect("ultimate", delimiter=",")
        writer = csv.DictWriter(out_csv, session_values["csv_fieldnames"].value)
        writer.writeheader()
        writer.writerows(rows)
        result = out_csv.getvalue()

    return result


def generate_xls_file_content(
    app: HanforFlask, filter_list: List[str] = None, invert_filter: bool = False
) -> io.BytesIO:
    """Generates the xlsx file content for a session."""
    from tags.tags import Tag

    requirements = get_requirements(app, filter_list=filter_list, invert_filter=invert_filter)
    var_collection = VariableCollection(app)
    tags = {tag.name: tag.internal for tag in app.db.get_objects(Tag).values()}

    # create  styles
    style_multiline = Alignment(vertical="top", wrap_text=True)
    style_bold = Font(bold=True)
    style_white = Font(color="FFFFFF", bold=True)
    style_filled = PatternFill(fill_type="solid", start_color="2a6ebb", end_color="2a6ebb")
    style_meta = PatternFill(fill_type="solid", start_color="004a99", end_color="004a99")
    # create excel template
    work_book = Workbook()
    work_sheet = work_book.active
    work_sheet.title = "Requirements"

    header_offset = 4

    def make_header(ws):
        ws.freeze_panes = "A4"
        for col in range(1, 10):
            for r in range(1, 3):
                ws.cell(r, col).fill = style_meta
        ws.cell(1, 2, value="HANFOR Report")
        ws.cell(1, 2).font = style_white
        ws.cell(1, 3, value=os.path.basename(app.db.get_object(SessionValue, "csv_input_file").value))
        ws.cell(1, 3).font = Font(color="FFFFFF")
        for col in range(1, 10):
            ws.cell(header_offset - 1, col).fill = style_filled
            ws.cell(header_offset - 1, col).font = style_white

    make_header(work_sheet)

    # Set column widths and headings
    work_sheet.column_dimensions["A"].width = 5
    work_sheet.cell(header_offset - 1, 1, value="Index")
    work_sheet.column_dimensions["B"].width = 20
    work_sheet.cell(header_offset - 1, 2, value="ID")
    work_sheet.column_dimensions["C"].width = 80
    work_sheet.cell(header_offset - 1, 3, value="Description")
    work_sheet.cell(header_offset - 1, 4, value="Type")
    work_sheet.column_dimensions["E"].width = 40
    work_sheet.cell(header_offset - 1, 5, value="Tags")
    work_sheet.cell(header_offset - 1, 6, value="Status")
    work_sheet.column_dimensions["G"].width = 160
    work_sheet.cell(header_offset - 1, 7, value="Formalisation")

    for i, requirement in enumerate(requirements):
        for c in range(1, 8):
            # Note: setting styles is ordering-sensitive so set styles FIRST
            work_sheet.cell(header_offset + i, c).alignment = style_multiline
        work_sheet.cell(header_offset + i, 1, requirement.pos_in_csv)
        work_sheet.cell(header_offset + i, 2).font = style_bold
        work_sheet.cell(header_offset + i, 2, requirement.rid)
        work_sheet.cell(header_offset + i, 3, requirement.description)
        work_sheet.cell(header_offset + i, 4, requirement.type_in_csv)

        work_sheet.cell(
            header_offset + i,
            5,
            "".join(
                [
                    f"{t}: {c} \n" if c else f"{t}\n"
                    for t, c in requirement.get_tag_name_comment_dict().items()
                    if t in tags and not tags[t]
                ]
            ),
        )
        work_sheet.cell(header_offset + i, 6, requirement.status)
        work_sheet.cell(header_offset + i, 7, "\n".join([f.get_string() for f in requirement.formalizations.values()]))

    # make severity sheet
    tag_sheet = work_book.create_sheet("Findings")
    make_header(tag_sheet)
    tag_sheet.column_dimensions["A"].width = 5
    tag_sheet.cell(header_offset - 1, 1, value="Index")
    tag_sheet.column_dimensions["B"].width = 20
    tag_sheet.cell(header_offset - 1, 2, value="ID")
    tag_sheet.cell(header_offset - 1, 3, value="Description")
    tag_sheet.column_dimensions["C"].width = 80
    tag_sheet.column_dimensions["D"].width = 20
    tag_sheet.cell(header_offset - 1, 4, value="Tag")
    tag_sheet.column_dimensions["E"].width = 60
    tag_sheet.cell(header_offset - 1, 5, value="Comment (Analysis)")
    tag_sheet.column_dimensions["F"].width = 10
    tag_sheet.cell(header_offset - 1, 6, value="Accept")
    tag_sheet.column_dimensions["G"].width = 15
    tag_sheet.cell(header_offset - 1, 7, value="Value")
    tag_sheet.column_dimensions["H"].width = 80
    tag_sheet.cell(header_offset - 1, 8, value="Comment (Review)")

    accept_state_validator = DataValidation(type="list", formula1='"TODO ,Accept,Decline,Inquery"', allow_blank=False)
    tag_sheet.add_data_validation(accept_state_validator)
    accept_state_validator.add("F4:F1048576")
    issue_value_validator = DataValidation(
        type="list", formula1='"TODO, 0 (no value),1 (nice to have),2 (useful),3 (possible desaster)"', allow_blank=True
    )
    tag_sheet.add_data_validation(issue_value_validator)
    issue_value_validator.add("G4:G1048576")

    issue_tags_reqs = []
    for req in requirements:
        for tag in req.get_tag_name_comment_dict():
            if tag in tags and tags[tag]:
                continue
            issue_tags_reqs.append((req, tag))

    for i, (req, tag) in enumerate(issue_tags_reqs):
        for c in range(1, 8):
            tag_sheet.cell(header_offset + i, c).alignment = style_multiline
        tag_sheet.cell(header_offset + i, 1, req.pos_in_csv)
        tag_sheet.cell(header_offset + i, 2).font = style_bold
        tag_sheet.cell(header_offset + i, 2, req.rid)
        tag_sheet.cell(header_offset + i, 3, req.description)
        tag_sheet.cell(header_offset + i, 4, tag)
        tag_sheet.cell(
            header_offset + i, 5, req.get_tag_name_comment_dict()[tag]
        )  # Tags do currently not have comments
        tag_sheet.cell(header_offset + i, 6, "TODO")
        tag_sheet.cell(header_offset + i, 7, "TODO")

    # make sheet with variables
    var_sheet = work_book.create_sheet("Variables")
    make_header(var_sheet)
    var_sheet.column_dimensions["A"].width = 40
    var_sheet.cell(header_offset - 1, 1, value="Name")
    var_sheet.column_dimensions["B"].width = 80
    var_sheet.column_dimensions["C"].width = 5
    var_sheet.cell(header_offset - 1, 2, value="Type")
    var_sheet.column_dimensions["D"].width = 180
    var_sheet.cell(header_offset - 1, 4, value="Invarianten")

    for i, var in enumerate(var_collection.collection.values()):
        for c in range(1, 8):
            var_sheet.cell(header_offset + i, c).alignment = style_multiline
        var_sheet.cell(header_offset + i, 1, var.name)
        var_sheet.cell(header_offset + i, 1).font = style_bold
        var_sheet.cell(header_offset + i, 2, var.type)
        var_sheet.cell(header_offset + i, 3, "E" if var.belongs_to_enum else "")
        var_sheet.cell(header_offset + i, 4, "\n".join([c.get_string() for c in var.get_constraints().values()]))

    work_book.active = tag_sheet
    buffer = io.BytesIO()
    work_book.save(buffer)
    return buffer


def clean_identifier_for_ultimate_parser(req_id: str, formalisation_id: int, used_identifiers: Set[str]) -> str:
    """Clean slug to be sound for ultimate parser.

    :param req_id: The requirement id to be cleaned.
    :param formalisation_id: The formalisation id to be cleaned.
    :param used_identifiers: Set of already used identifiers.
    :return: (identifier, used_slugs) save_slug a save to use form of slug. save_slug added to used_slugs.
    """
    # Replace any occurrence of [whitespace, `.`, `-`] with `_`
    base_identifier = re.sub(r"[\s+.-]+", "_", req_id.strip())

    # Resolve illegal start by prepending the slug with ID_ in case it does not start with a letter.
    base_identifier = re.sub(r"^([^a-zA-Z])", r"ID_\1", base_identifier)

    def create_identifier(base: str, extension: int, base_suffix: int = -1):
        if base_suffix == -1:
            return "{}_{}".format(base, extension)
        return "{}_{}_{}".format(base, base_suffix, extension)

    identifier = create_identifier(base_identifier, formalisation_id)

    # Resolve duplicates
    # search for the first free suffix.
    if identifier in used_identifiers:
        suffix = 1

        while create_identifier(base_identifier, formalisation_id, suffix) in used_identifiers:
            suffix += 1
        identifier = create_identifier(base_identifier, formalisation_id, suffix)

    used_identifiers.add(identifier)

    return identifier


def generate_req_file_content(app: HanforFlask, filter_list=None, invert_filter=False, variables_only=False):
    """Generate the content (string) for the ultimate requirements file.
    :param app: Current app.
    :type app: FlaskApp
    :param filter_list: A list of requirement IDs to be included in the result. All if not set.
    :type filter_list: list (of strings)
    :param invert_filter: Exclude filter
    :type invert_filter: bool
    :type invert_filter: bool
    :param variables_only: If true, only variables and no requirements will be included.
    :return: Content for the req file.
    :rtype: str
    """
    logging.info("Generating .req file content for session {}".format(app.config["SESSION_TAG"]))
    # Get requirements
    requirements = get_requirements(app, filter_list=filter_list, invert_filter=invert_filter)

    var_collection = VariableCollection(app)
    available_vars = []
    if filter_list is not None:
        # Filter the available vars to only include the ones actually used by a requirement.
        logging.info("Filtering the .req file to only include {} selected requirements.".format(len(filter_list)))
        logging.debug("Only include req.ids: {}.".format(filter_list))
        target_set = set(filter_list)
        for var in var_collection.get_available_vars_list(sort_by="name"):
            try:
                used_in = var_collection.var_req_mapping[var["name"]]
                if used_in & target_set:
                    available_vars.append(var)
            except Exception:  # noqa
                logging.debug("Ignoring variable `{}`".format(var))
    else:
        available_vars = var_collection.get_available_vars_list(sort_by="name")

    available_vars = [var["name"] for var in available_vars]

    content_list = []
    constants_list = []
    constraints_list = []

    # Parse variables and variable constraints.
    for var in var_collection.collection.values():
        if var.name in available_vars:
            if var.type in ["CONST", "ENUMERATOR_INT", "ENUMERATOR_REAL"]:
                constants_list.append("CONST {} IS {}".format(var.name, var.value))
            else:
                content_list.append(
                    "Input {} IS {}".format(var.name, boogie_parsing.BoogieType.reverse_alias(var.type).name)
                )
            try:
                for index, constraint in var.constraints.items():
                    if constraint.scoped_pattern is None:
                        continue
                    if constraint.scoped_pattern.get_scope_slug().lower() == "none":
                        continue
                    if constraint.scoped_pattern.get_pattern_slug() in ["NotFormalizable", "None"]:
                        continue
                    if len(constraint.get_string()) == 0:
                        continue
                    constraints_list.append(
                        "Constraint_{}_{}: {}".format(re.sub(r"\s+", "_", var.name), index, constraint.get_string())
                    )
            except Exception:
                # TODO: this is not a nice way to solve this
                pass
    content_list.sort()
    constants_list.sort()
    constraints_list.sort()

    content = "\n".join(content_list)
    constants = "\n".join(constants_list)
    constraints = "\n".join(constraints_list)

    if len(constants) > 0:
        content = "\n\n".join([constants, content])
    if len(constraints) > 0:
        content = "\n\n".join([content, constraints])
    content += "\n\n"

    # parse requirement formalizations.
    if not variables_only:
        used_identifiers = set()
        for requirement in requirements:  # type: Requirement
            try:
                for index, formalization in requirement.formalizations.items():
                    identifier = clean_identifier_for_ultimate_parser(requirement.rid, index, used_identifiers)
                    if formalization.scoped_pattern is None:
                        continue
                    if formalization.scoped_pattern.get_scope_slug().lower() == "none":
                        continue
                    if formalization.scoped_pattern.get_pattern_slug() in ["NotFormalizable", "None"]:
                        continue
                    if len(formalization.get_string()) == 0:
                        # Formalization string is empty if expressions are missing or none set. Ignore in output
                        continue
                    content += "{}: {}\n".format(identifier, formalization.get_string())
            except AttributeError:
                continue
    content += "\n"

    return content


def get_stored_session_names(session_folder, only_names=False, with_revisions=False) -> tuple:
    """Get stored session tags (folder names) including os.stat.
    Returned tuple is (
        (os.stat(), name),
        ...
    )
    If only_names == True the tuple is (
        name_1,
        ...
    )
    If with_with_revisions == True the tuple is (
        {
            name: 'name_1',
            'revisions': [revision_1, ...],
            revisions_stats: {
                revision_1: {
                    name: 'revision_1',
                    last_mod: "%A %d. %B %Y at %X" formatted mtime,
                    num_vars: Number of variables used in this revision.
                }
            }
        },
        ...
    )

    Args:
        session_folder (str): path to hanfor data folder.
        only_names (bool): Only return the names (tags).
        with_revisions (bool): Recurse also into the available revisions.

    Returns:
        tuple: Only folder names or stats with names.
    """
    result = ()
    if not session_folder:
        session_folder = os.path.join(here, "data")

    try:
        result = [
            (os.path.join(session_folder, file_name), file_name)
            for file_name in os.listdir(session_folder)
            if os.path.isdir(os.path.join(session_folder, file_name))
        ]
        if with_revisions:
            result = (
                {
                    "name": entry[1],
                    "revisions": get_available_revisions(None, entry[0]),
                    "revisions_stats": get_revisions_with_stats(entry[0]),
                }
                for entry in result
            )
        elif only_names:
            result = (entry[1] for entry in result)
        else:
            result = ((os.stat(entry[0]), entry[1]) for entry in result)
    except Exception as e:
        logging.error("Could not fetch stored sessions: {}".format(e))

    return result


def get_revisions_with_stats(session_path):
    """Get meta information about available revisions for a given session path.

    Returns a dict with revision name as key for each revision.
    Each item is then a dict like:
        {
            name: 'revision_1',
            last_mod: %A %d. %B %Y at %X formatted mtime,
            num_vars: 9001
        }


    :param session_path: { revision_1: { name: 'revision_1', last_mod: %A %d. %B %Y at %X, num_vars: 9001} ... }
    """
    revisions = get_available_revisions(None, session_path)
    revisions_stats = dict()
    for revision in revisions:
        revision_path = os.path.join(session_path, revision)
        tmp_db = JsonDatabase(read_only=True)
        add_custom_serializer_to_database(tmp_db)
        tmp_db.init_tables(revision_path)

        revisions_stats[revision] = {
            "name": revision,
            "last_mod": get_last_edit_from_path(revision_path),
            "num_vars": len(tmp_db.get_objects(Variable)),
        }
    return revisions_stats


def get_last_edit_from_path(path_str):
    """Return a human readable form of the last edit (mtime) for a path.

    :param path_str: str to path.
    :return: "%A %d. %B %Y at %X" formatted mtime
    """
    return datetime.datetime.fromtimestamp(os.stat(path_str).st_mtime).strftime("%A %d. %B %Y at %X")


def get_available_revisions(config, folder=None):
    """Returns available revisions for a given session folder.
    Uses config['SESSION_FOLDER'] if no explicit folder is given.

    :param config:
    :param folder:
    :return:
    """
    result = []

    if folder is None:
        folder = config["SESSION_FOLDER"]

    try:
        names = os.listdir(folder)
        result = [name for name in names if os.path.isdir(os.path.join(folder, name))]
    except Exception as e:
        logging.error("Could not fetch stored revisions: {}".format(e))

    return result


def enable_logging(log_level=logging.ERROR, to_file=False, filename="reqtransformer.log"):
    """Enable Logging.

    :param log_level: Log level
    :type log_level: int
    :param to_file: Whether output should be stored in a file.
    :type to_file: bool
    :param filename: Filename to log to.
    :type filename: str
    """
    if to_file:
        logging.basicConfig(format="%(asctime)s: [%(levelname)s]: %(message)s", filename=filename, level=log_level)
    else:
        logging.basicConfig(format="%(asctime)s: [%(levelname)s]: %(message)s", level=log_level)
    logging.debug("Enabled logging.")


def setup_logging(app: HanforFlask):
    """Initializes logging with settings from the config."""
    if app.config["LOG_LEVEL"] == "DEBUG":
        log_level = logging.DEBUG
    elif app.config["LOG_LEVEL"] == "INFO":
        log_level = logging.INFO
    elif app.config["LOG_LEVEL"] == "WARNING":
        log_level = logging.WARNING
    else:
        log_level = logging.ERROR

    enable_logging(log_level=log_level, to_file=app.config["LOG_TO_FILE"], filename=app.config["LOG_FILE"])


def register_assets(app: HanforFlask):
    # TODO: Unused functionality.
    bundles = {}

    assets = Environment(app)
    assets.register(bundles)


def get_datatable_additional_cols(app: HanforFlask):
    offset = 8  # we have 8 fixed cols.
    result = list()

    for index, name in enumerate(app.db.get_object(SessionValue, "csv_fieldnames").value):
        result.append(
            {
                "target": index + offset,
                "csv_name": "csv_data.{}".format(name),
                "table_header_name": "csv: {}".format(name),
            }
        )

    return {"col_defs": result}


def add_msg_to_flask_session_log(app: HanforFlask, message: str, req_list: list[Requirement] = None) -> None:
    """Add a log message for the frontend_logs.

    :param req_list: A list of affected requirements
    :param app: The flask app
    :param message: Log message string
    """
    app.db.add_object(RequirementEditHistory(datetime.datetime.now(), message, req_list))


def get_flask_session_log(app: HanforFlask, html_format: bool = False) -> Union[list, str]:
    """Get the frontend log messages from frontend_logs.

    :param app: The flask app
    :param html_format: Return formatted html version.
    :return: list of messages or html string in case `html == True`
    """
    history_elements = app.db.get_objects(RequirementEditHistory)

    if html_format:
        result = ""
        for element in history_elements.values():  # type: RequirementEditHistory
            links = ",".join(
                [
                    ' <a class="req_direct_link" href="#" data-rid="{rid}">{rid}</a>'.format(rid=req.rid)
                    for req in element.requirements
                ]
            )
            result += f"<p>[{element.timestamp}] {element.message}{links}</p>"
    else:
        result = [history_elements.items()]

    return result


def slugify(s):
    """Normalizes string, converts to lowercase, removes non-alpha characters, and converts spaces to hyphens.

    :param s: string
    :type s: str
    :return: String save for filename
    :rtype: str
    """
    s = str(s).strip().replace(" ", "_")
    return re.sub(r"(?u)[^-\w.]", "", s)


def generate_file_response(content, name, mimetype="text/plain"):
    response = Response(content, mimetype=mimetype)
    response.headers["Content-Disposition"] = f"attachment; filename*=UTF-8''{name}"
    return response


class PrefixMiddleware(object):
    """Support for url prefixes."""

    def __init__(self, app, prefix=""):
        self.app = app
        self.prefix = prefix

    def __call__(self, environ, start_response):
        if environ["PATH_INFO"].startswith(self.prefix):
            environ["PATH_INFO"] = environ["PATH_INFO"][len(self.prefix) :]
            environ["SCRIPT_NAME"] = self.prefix
            return self.app(environ, start_response)
        else:
            start_response("404", [("Content-Type", "text/plain")])
            return ["Sorry, this url does not belong to Hanfor.".encode()]


class ListStoredSessions(argparse.Action):
    """List available session tags."""

    def __init__(self, option_strings, app, dest, *args, **kwargs):
        self.app = app
        super(ListStoredSessions, self).__init__(option_strings=option_strings, dest=dest, *args, **kwargs)

    def __call__(self, *args, **kwargs):
        entries = get_stored_session_names(self.app.config["SESSION_BASE_FOLDER"], with_revisions=True)
        data = [["Tag", "Revision", "Last Modified"]]
        for entry in entries:
            revisions = list()
            for name, values in entry["revisions_stats"].items():
                revisions.append((name, values["last_mod"]))
            revisions.sort()
            if len(revisions) > 0:
                data.append([entry["name"], revisions[0][0], revisions[0][1]])
                for i in range(1, len(revisions)):
                    data.append(["", revisions[i][0], revisions[i][1]])
            else:
                data.append([entry["name"], "", ""])
        print("Stored sessions: ")
        if len(data) > 1:
            print(DoubleTable(data).table)
        else:
            print("No sessions in found.")
        exit(0)


class GenerateScopedPatternTrainingData(argparse.Action):
    """Generate training data consisting of requirement descriptions with assigned scoped pattern."""

    def __init__(self, option_strings, app, dest, *args, **kwargs):
        self.app = app
        super(GenerateScopedPatternTrainingData, self).__init__(
            option_strings=option_strings, dest=dest, *args, **kwargs
        )

    def __call__(self, *args, **kwargs):
        # logging.debug(self.app.config)
        entries = get_stored_session_names(self.app.config["SESSION_BASE_FOLDER"])
        result = dict()
        for entry in entries:
            logging.debug("Looking into {}".format(entry[1]))
            current_session_folder = os.path.join(self.app.config["SESSION_BASE_FOLDER"], entry[1])
            revisions = get_available_revisions(self.app.config, folder=current_session_folder)
            for revision in revisions:
                current_revision_folder = os.path.join(str(current_session_folder), revision)
                logging.debug("Processing `{}`".format(current_revision_folder))
                requirements = get_requirements(self.app)
                logging.debug("Found {} requirements .. fetching the formalized ones.".format(len(requirements)))
                used_slugs = set()
                for requirement in requirements:
                    try:
                        if len(requirement.description) == 0:
                            continue
                        slug, used_slugs = clean_identifier_for_ultimate_parser(requirement.rid, used_slugs)
                        result[slug] = dict()
                        result[slug]["desc"] = requirement.description
                        for index, formalization in requirement.formalizations.items():
                            if formalization.scoped_pattern is None:
                                continue
                            if formalization.scoped_pattern.get_scope_slug().lower() == "none":
                                continue
                            if formalization.scoped_pattern.get_pattern_slug() in ["NotFormalizable", "None"]:
                                continue
                            if len(formalization.get_string()) == 0:
                                # formalization string is empty if expressions are missing or none set. Ignore in output
                                continue
                            f_key = "formalization_{}".format(index)
                            result[slug][f_key] = dict()
                            result[slug][f_key]["scope"] = formalization.scoped_pattern.get_scope_slug()
                            result[slug][f_key]["pattern"] = formalization.scoped_pattern.get_pattern_slug()
                            result[slug][f_key]["formalization"] = formalization.get_string()
                    except AttributeError:
                        continue
            with open("training_data.json", mode="w", encoding="utf-8") as f:
                json.dump(result, f)
        exit(0)


class HanforArgumentParser(argparse.ArgumentParser):
    def __init__(self, app):
        super().__init__()
        self.app = app
        self.add_argument("tag", help="A tag for the session. Session will be reloaded, if tag exists.")
        self.add_argument("-c", "--input_csv", help="Path to the csv to be processed.")
        self.add_argument(
            "-r",
            "--revision",
            action="store_true",
            help="Create a new session by updating a existing session with a new csv file.",
        )
        self.add_argument(
            "-rti", "--reload_type_inference", action="store_true", help="Reload the type inference results."
        )
        self.add_argument(
            "-L",
            "--list_stored_sessions",
            nargs=0,
            help="List the tags of stored sessions..",
            action=ListStoredSessions,
            app=self.app,
        )
        self.add_argument(
            "-G",
            "--generate_scoped_pattern_training_data",
            nargs=0,
            help="Generate training data out of description with assigned scoped pattern.",
            action=GenerateScopedPatternTrainingData,
            app=self.app,
        )
        self.add_argument(
            "-hd",
            "--headers",
            type=str,
            help='Header Definition of the form --header=\'{ "csv_id_header": "ID", "csv_desc_header": "Description", '
            '"csv_formal_header": "Hanfor_Formalization", "csv_type_header" : "Type"}\', must be valid json.',
            default=None,
        )


class Revision:
    def __init__(self, app: HanforFlask, args, base_revision_name):
        self.app = app
        self.args = args
        self.base_revision_name = base_revision_name
        self.is_initial_revision = True
        self.base_revision_folder = None
        self.base_revision_db = None
        self.available_sessions = None
        self.requirement_collection = None

        self._set_is_initial_revision()
        self._set_revision_name()
        self._set_base_revision_folder()

    def create_revision(self):
        self._check_base_revision_available()
        self._set_config_vars()
        self._set_available_sessions()
        if self.is_initial_revision:
            self._create_revision_folder()
        else:
            self._copy_base_revision()

        self._set_base_revision_db()
        self.app.db.init_tables(self.app.config["REVISION_FOLDER"])
        self._try_save(self._load_from_csv, "Could not read CSV")
        if self.is_initial_revision:
            for req in self.requirement_collection.requirements:
                self.app.db.add_object(req, delay_update=True)
            self.app.db.update()
            self._generate_session_values()
        else:
            self._update_session_values()
            self._try_save(self._merge_with_base_revision, " Could not merge with base session")

    def _try_save(self, what, error_msg):
        """Safely run a method. Cleanup the revision in case of an exception.

        Args:
            what: Method to run.
            error_msg: Message in case of failure.
        """
        try:
            what()
        except Exception as e:
            logging.error("Abort creating revision. {msg}: {error}".format(msg=error_msg, error=type(e)))
            self._revert_and_cleanup()
            raise e

    def _set_revision_name(self):
        if not self.base_revision_name:
            logging.info("No revisions for `{}`. Creating initial revision.".format(self.args.tag))
            self.revision_name = "revision_0"
        # Revision based on an existing revision
        else:
            new_revision_count = max([int(name.split("_")[1]) for name in get_available_revisions(self.app.config)]) + 1
            self.revision_name = "revision_{}".format(new_revision_count)

    def _set_base_revision_db(self):
        if not self.is_initial_revision:
            self.base_revision_db = JsonDatabase(read_only=True)
            add_custom_serializer_to_database(self.base_revision_db)
            self.base_revision_db.init_tables(self.base_revision_folder)

    def _set_base_revision_folder(self):
        if not self.is_initial_revision:
            self.base_revision_folder = os.path.join(self.app.config["SESSION_FOLDER"], self.base_revision_name)

    def _set_is_initial_revision(self):
        if self.base_revision_name:
            self.is_initial_revision = False

    def _set_config_vars(self):
        self.app.config["USING_REVISION"] = self.revision_name
        self.app.config["REVISION_FOLDER"] = os.path.join(self.app.config["SESSION_FOLDER"], self.revision_name)

    def _check_base_revision_available(self):
        if not self.is_initial_revision and self.base_revision_name not in get_available_revisions(self.app.config):
            logging.error(
                "Base revision `{}` not found in `{}`".format(
                    self.base_revision_name, self.app.config["SESSION_FOLDER"]
                )
            )
            raise FileNotFoundError

    def _set_available_sessions(self):
        self.available_sessions = get_stored_session_names(self.app.config["SESSION_BASE_FOLDER"], with_revisions=True)

    def _load_from_csv(self):
        # Load requirements from .csv file and store them into separate requirements.
        if self.args.input_csv is None:
            HanforArgumentParser(self.app).error("Creating an (initial) revision requires -c INPUT_CSV")
        self.requirement_collection = RequirementCollection()
        base_revision_headers = {}
        if self.base_revision_db:
            base_revision_headers = self.base_revision_db.get_objects(SessionValue)
        self.requirement_collection.create_from_csv(
            csv_file=self.args.input_csv,
            app=self.app,
            input_encoding="utf8",
            base_revision_headers=base_revision_headers,
            user_provided_headers=(json.loads(self.args.headers) if self.args.headers else None),
            available_sessions=self.available_sessions,
        )

    def _create_revision_folder(self):
        os.makedirs(self.app.config["REVISION_FOLDER"], exist_ok=True)

    def _copy_base_revision(self):
        copytree(self.base_revision_folder, self.app.config["REVISION_FOLDER"])

    def _revert_and_cleanup(self):
        logging.info("Reverting revision creation.")
        if self.is_initial_revision:
            logging.debug("Revert initialized session folder. Deleting: `{}`".format(self.app.config["SESSION_FOLDER"]))
            remove_json_database_data_tracing_logger(True)
            shutil.rmtree(self.app.config["SESSION_FOLDER"])
        else:
            logging.debug(
                "Revert initialized revision folder. Deleting: `{}`".format(self.app.config["REVISION_FOLDER"])
            )
            remove_json_database_data_tracing_logger(True)
            shutil.rmtree(self.app.config["REVISION_FOLDER"])

    def _generate_session_values(self):
        # Generate the session values: Store some meta information.
        self.app.db.add_object(SessionValue("csv_input_file", self.args.input_csv))
        self.app.db.add_object(SessionValue("csv_fieldnames", self.requirement_collection.csv_meta.fieldnames))
        self.app.db.add_object(SessionValue("csv_id_header", self.requirement_collection.csv_meta.id_header))
        self.app.db.add_object(SessionValue("csv_formal_header", self.requirement_collection.csv_meta.formal_header))
        self.app.db.add_object(SessionValue("csv_type_header", self.requirement_collection.csv_meta.type_header))
        self.app.db.add_object(SessionValue("csv_desc_header", self.requirement_collection.csv_meta.desc_header))
        self.app.db.add_object(SessionValue("csv_hash", hash_file_sha1(self.args.input_csv)))

    def _update_session_values(self):
        # Update the session values: Store some meta information.
        self.app.db.get_object(SessionValue, "csv_input_file").value = self.args.input_csv
        self.app.db.get_object(SessionValue, "csv_fieldnames").value = self.requirement_collection.csv_meta.fieldnames
        self.app.db.get_object(SessionValue, "csv_id_header").value = self.requirement_collection.csv_meta.id_header
        self.app.db.get_object(SessionValue, "csv_formal_header").value = (
            self.requirement_collection.csv_meta.formal_header
        )
        self.app.db.get_object(SessionValue, "csv_type_header").value = self.requirement_collection.csv_meta.type_header
        self.app.db.get_object(SessionValue, "csv_desc_header").value = self.requirement_collection.csv_meta.desc_header
        self.app.db.get_object(SessionValue, "csv_hash").value = hash_file_sha1(self.args.input_csv)

    def _merge_with_base_revision(self):
        # Merge the old revision into the new revision
        logging.info(f"Merging `{self.base_revision_name}` into `{self.revision_name}`.")
        reqs: dict[str, Requirement] = dict(self.app.db.get_objects(Requirement))
        new_reqs: dict[str, Requirement] = {r.rid: r for r in self.requirement_collection.requirements}

        req_ids: set[str] = set(reqs.keys())
        new_req_ids: set[str] = set(new_reqs.keys())

        # delete deleted requirements
        for rid in req_ids.difference(new_req_ids):
            self.app.db.remove_object(reqs[rid])

        # insert and tag new reqs
        revision_tag_name = f"{self.base_revision_name}_to_{self.revision_name}_new_requirement"
        if not self.app.db.key_in_table(Tag, revision_tag_name):
            revision_tag = Tag(revision_tag_name, Color.BS_INFO.value, False, "")
        else:
            revision_tag = self.app.db.get_object(Tag, revision_tag_name)

        for rid in new_req_ids.difference(req_ids):
            logging.info("Add newly introduced requirement `{}`".format(rid))
            new_reqs[rid].tags[revision_tag] = ""
            self.app.db.add_object(new_reqs[rid])

        # updating existing requirements
        data_changed_tag_name = f"{self.base_revision_name}_to_{self.revision_name}_data_changed"
        if not self.app.db.key_in_table(Tag, data_changed_tag_name):
            data_changed_tag = Tag(data_changed_tag_name, Color.BS_INFO.value, False, "")
        else:
            data_changed_tag = self.app.db.get_object(Tag, data_changed_tag_name)

        description_changed_tag_name = f"{self.base_revision_name}_to_{self.revision_name}_description_changed"
        if not self.app.db.key_in_table(Tag, description_changed_tag_name):
            description_changed_tag = Tag(description_changed_tag_name, Color.BS_INFO.value, False, "")
        else:
            description_changed_tag = self.app.db.get_object(Tag, description_changed_tag_name)

        migrated_formalization_tag_name = f"{self.base_revision_name}_to_{self.revision_name}_migrated_formalization"
        if not self.app.db.key_in_table(Tag, migrated_formalization_tag_name):
            migrated_formalization_tag = Tag(migrated_formalization_tag_name, Color.BS_INFO.value, False, "")
        else:
            migrated_formalization_tag = self.app.db.get_object(Tag, migrated_formalization_tag_name)

        for rid in new_req_ids.intersection(req_ids):
            r = reqs[rid]
            new_r = new_reqs[rid]
            # add revision diff and update description, type_in_csv, csv_row and pos_in_csv
            if r.description != new_r.description:
                logging.info(f"Description changed. Add `description_changed` tag to `{rid}`.")
                r.tags[description_changed_tag] = ""
                r.status = "Todo"

            r.revision_diff = new_r

            if len(r.revision_diff) > 0:
                logging.info(f"CSV entry changed. Add `revision_data_changed` tag to `{rid}`.")
                r.tags[data_changed_tag] = ""

            # If the new formalization is empty: just migrate the formalization.
            #  - Tag with `migrated_formalization` if the description changed.
            if len(new_r.formalizations) == 0 and len(r.formalizations) == 0:
                pass
            elif len(new_r.formalizations) == 0 and len(r.formalizations) > 0:
                logging.info("Migrate formalization for `{}`".format(rid))
                if r.description != new_r.description:
                    logging.info(
                        "Add `migrated_formalization` tag to `{}`, status to `Todo` since description changed".format(
                            rid
                        )
                    )
                    r.tags[migrated_formalization_tag] = ""
                    r.status = "Todo"
            elif len(new_r.formalizations) > 0 and len(r.formalizations) == 0:
                logging.error("Parsing of the requirement not supported.")
                raise NotImplementedError
            else:
                logging.error("Parsing of the requirement not supported.")
                raise NotImplementedError

        # Store the updated requirements for the new revision.
        logging.info("Store merge changes to revision `{}`".format(self.revision_name))
        self.app.db.update()


def add_custom_serializer_to_database(database: JsonDatabase) -> None:

    def scope_serialize(obj: Scope, db_serializer: Callable[[any, str], any], user: str) -> dict:
        return db_serializer(obj.value, user)

    def scope_deserialize(data: dict, db_deserializer: Callable[[any], any]) -> Scope:
        return Scope(db_deserializer(data))

    database.add_custom_serializer(Scope, scope_serialize, scope_deserialize)

    def datetime_serialize(obj: datetime.datetime, db_serializer: Callable[[any, str], any], user: str) -> dict:
        return db_serializer(obj.isoformat(), user)

    def datetime_deserialize(data: dict, db_deserializer: Callable[[any], any]) -> datetime.datetime:
        return datetime.datetime.fromisoformat(db_deserializer(data))

    database.add_custom_serializer(datetime.datetime, datetime_serialize, datetime_deserialize)

    def boogie_type_serialize(
        obj: boogie_parsing.BoogieType, db_serializer: Callable[[any, str], any], user: str
    ) -> dict:
        return db_serializer(obj.value, user)

    def boogie_type_deserialize(data: dict, db_deserializer: Callable[[any], any]) -> boogie_parsing.BoogieType:
        return boogie_parsing.BoogieType(db_deserializer(data))

    database.add_custom_serializer(boogie_parsing.BoogieType, boogie_type_serialize, boogie_type_deserialize)
